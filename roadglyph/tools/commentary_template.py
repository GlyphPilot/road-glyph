import os
import json
import gzip
import re
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

# commentary root
ROOT_DIR = "/path/to/dataset/commentary/simlingo/training_3_scenarios/routes_training/random_weather_seed_3_balanced_100"

# measurements root (data side)
DATA_ROOT = "/path/to/dataset/data/simlingo/training_3_scenarios/routes_training/random_weather_seed_3_balanced_100"

OUT_XLSX = "simlingo_commentary_templates_training_3_scenarios.xlsx"

EXCEL_CELL_MAX = 32767
TRUNCATE_TO = 32000

def read_json_gz(path: str):
    with gzip.open(path, "rt", encoding="utf-8") as f:
        return json.load(f)

def safe_cell_text(x):
    if x is None:
        return ""
    if not isinstance(x, str):
        x = json.dumps(x, ensure_ascii=False)
    if len(x) > EXCEL_CELL_MAX:
        x = x[:TRUNCATE_TO] + " ...[TRUNCATED]"
    return x

# -----------------------------
# Rule-based template classification
# -----------------------------
def classify_route_action(t: str) -> str:
    if not t:
        return ""
    s = t.lower()

    # (3) end of deviation
    if "return to your original route" in s or "return to your original lane" in s:
        return "DEVIATION_END"

    # (2) during deviation: prefix is the key indicator
    if "stay on your current lane to" in s:
        return "DEVIATION_DURING"

    # (special) waiting for a lane change gap before deviation
    if "wait for a gap in the traffic" in s:
        return "DEVIATION_BEFORE_WAIT_GAP"

    # (1) before deviation: scenario-specific action
    if "overtake" in s:
        return "DEVIATION_BEFORE_OVERTAKE"
    if "give way" in s:
        return "DEVIATION_BEFORE_GIVE_WAY"
    if "go around" in s or "circumvent" in s or "avoid" in s:
        return "DEVIATION_BEFORE_GO_AROUND"

    # default
    if s.startswith("follow the route"):
        return "FOLLOW_ROUTE"
    return "OTHER"

def classify_speed_action(t: str) -> str:
    if not t:
        return ""
    s = t.lower()

    if "remain stopped" in s:
        return "REMAIN_STOPPED"
    if "come to a stop now" in s:
        return "STOP_NOW"
    if "maintain your current speed" in s:
        return "MAINTAIN_SPEED"
    if "maintain the reduced speed" in s:
        return "MAINTAIN_REDUCED_SPEED"
    if "increase your speed" in s or "accelerate" in s:
        return "INCREASE_SPEED"
    if "slow down" in s or "decelerate" in s:
        return "SLOW_DOWN"
    if re.search(r"\bstop\b", s):
        return "STOP"
    return "UNKNOWN"

def classify_speed_reason(t: str) -> str:
    if not t:
        return ""
    s = t.lower()

    if "pedestrian" in s or "child" in s:
        return "PEDESTRIAN"
    if "traffic light" in s:
        if "red traffic light" in s:
            return "TRAFFIC_LIGHT_RED"
        if "green" in s:
            return "TRAFFIC_LIGHT_GREEN"
        return "TRAFFIC_LIGHT"
    if "stop sign" in s:
        return "STOP_SIGN"
    if "junction" in s:
        return "JUNCTION_NOTICE"
    if "construction" in s or "cones" in s:
        return "CONSTRUCTION"
    if "accident" in s:
        return "ACCIDENT"
    if "emergency vehicle" in s:
        return "EMERGENCY_VEHICLE"
    if "vehicle" in s or "car" in s or "<object>" in s:
        if "stay behind" in s or "remain behind" in s or "follow" in s:
            return "LEAD_VEHICLE_FOLLOW"
        if "avoid a collision" in s or "to avoid a collision" in s:
            return "COLLISION_AVOIDANCE"
        return "VEHICLE"
    if "to reach the target speed" in s:
        return "REACH_TARGET_SPEED"
    return "NONE_OR_UNKNOWN"

# -----------------------------
# commentary rel_path -> measurements path conversion
# -----------------------------
def rel_commentary_to_measurements(rel_path: str) -> str:
    # e.g. Town.../commentary/0003.json.gz -> Town.../measurements/0003.json.gz
    return rel_path.replace(f"{os.sep}commentary{os.sep}", f"{os.sep}measurements{os.sep}")

def load_hlc(meas_path: str):
    """
    returns (command, next_command) or ("", "") if missing
    """
    try:
        m = read_json_gz(meas_path)
        if isinstance(m, dict):
            return m.get("command", ""), m.get("next_command", "")
        return "", ""
    except (FileNotFoundError, json.JSONDecodeError, OSError):
        return "", ""

# Optional: modify HLC_MAP to customize the human-readable command string mapping
HLC_MAP = {
    1: "LEFT",
    2: "RIGHT",
    3: "STRAIGHT",
    4: "LANEFOLLOW",
    5: "CHANGE_LANE_LEFT",
    6: "CHANGE_LANE_RIGHT",
}

def hlc_to_str(x):
    try:
        xi = int(x)
        return HLC_MAP.get(xi, str(xi))
    except Exception:
        return ""

def main():
    wb = Workbook()
    ws = wb.active
    ws.title = "templates"

    ws.append(["COMMENTARY_ROOT_DIR", ROOT_DIR])
    ws.append(["MEASUREMENTS_DATA_ROOT", DATA_ROOT])
    ws.append([])

    headers = [
        "relative_path_from_commentary_root",
        "relative_measurements_path",
        "commentary_template",
        "route_action",
        "speed_action",
        "speed_reason",
        "command",
        "command_str",
        "next_command",
        "next_command_str",
    ]
    ws.append(headers)

    n_total = n_written = n_missing = n_error = 0
    n_meas_missing = 0

    # --- (A) collect file list ---
    all_files = []
    for dirpath, _, filenames in os.walk(ROOT_DIR):
        for fn in filenames:
            if fn.endswith(".json.gz") or fn.endswith(".json.jz"):
                fpath = os.path.join(dirpath, fn)
                rel_path = os.path.relpath(fpath, ROOT_DIR)
                all_files.append((rel_path, fpath))

    # --- (B) sort key: (route folder, frame index) ---
    def sort_key(item):
        rel_path, _ = item
        base = os.path.basename(rel_path)  # ex) "0010.json.gz"
        m = re.search(r"(\d+)", base)
        frame_idx = int(m.group(1)) if m else -1
        folder = os.path.dirname(rel_path)  # ex) "Town.../commentary"
        return (folder, frame_idx, base)

    all_files.sort(key=sort_key)

    # --- (C) process in sorted order ---
    for rel_path, fpath in all_files:
        n_total += 1
        try:
            data = read_json_gz(fpath)
            template = data.get("commentary_template", "") if isinstance(data, dict) else ""
            if template == "":
                n_missing += 1

            # classify
            route_action = classify_route_action(template)
            speed_action = classify_speed_action(template)
            speed_reason = classify_speed_reason(template)

            # compute measurements path and load HLC
            meas_rel = rel_commentary_to_measurements(rel_path)
            meas_path = os.path.join(DATA_ROOT, meas_rel)
            command, next_command = load_hlc(meas_path)
            if command == "" and next_command == "":
                n_meas_missing += 1

            ws.append([
                rel_path,
                meas_rel,
                safe_cell_text(template),
                route_action,
                speed_action,
                speed_reason,
                command,
                hlc_to_str(command),
                next_command,
                hlc_to_str(next_command),
            ])
            n_written += 1

            if n_total % 5000 == 0:
                print(f"[progress] scanned={n_total:,} written={n_written:,} missing_template={n_missing:,} meas_missing={n_meas_missing:,} error={n_error:,}")

        except Exception as e:
            n_error += 1
            if n_error <= 20:
                print(f"[error] {fpath} -> {e}")
            elif n_error == 21:
                print("[error] too many errors... (suppressing further messages)")

    col_widths = [80, 80, 110, 18, 22, 28, 10, 18, 12, 18]
    for i, w in enumerate(col_widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.freeze_panes = "A5"
    wb.save(OUT_XLSX)

    print("\nDone!")
    print(f"commentary root: {ROOT_DIR}")
    print(f"data root      : {DATA_ROOT}")
    print(f"out            : {OUT_XLSX}")
    print(f"scanned={n_total:,}, written={n_written:,}, missing_template={n_missing:,}, meas_missing={n_meas_missing:,}, errors={n_error:,}")

if __name__ == "__main__":
    main()